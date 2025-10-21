
import os
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

MONTHS_UA = {
    1: 'Січень', 2: 'Лютий', 3: 'Березень', 4: 'Квітень', 5: 'Травень', 6: 'Червень',
    7: 'Липень', 8: 'Серпень', 9: 'Вересень', 10: 'Жовтень', 11: 'Листопад', 12: 'Грудень'
}

MONTHS_UA_ADJECTIVE = {
    1: 'Січневий', 2: 'Лютневий', 3: 'Березневий', 4: 'Квітневий', 5: 'Травневий', 6: 'Червневий',
    7: 'Липневий', 8: 'Серпневий', 9: 'Вересневий', 10: 'Жовтневий', 11: 'Листопадовий', 12: 'Грудневий'
}

def get_report_filename(reports_dir: str, for_date: datetime = None) -> str:
    """Генерує назву файлу звіту для місяця, що відповідає даті for_date."""
    if for_date is None:
        for_date = datetime.now()
    
    month_name_adj = MONTHS_UA_ADJECTIVE[for_date.month]
    filename = f"{month_name_adj}_звіт_заміни.xlsx"
    
    os.makedirs(reports_dir, exist_ok=True)
    return os.path.join(reports_dir, filename)

def style_header_cell(cell):
    """Стилізує клітинку заголовка."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border

def record_replacement_to_excel(reports_dir: str, replacement_data: dict):
    """Записує дані про заміну в Excel-файл."""
    filepath = get_report_filename(reports_dir)
    headers = [
        "ID Заявки", "Керівник", "Дата Заміни", "Посада",
        "Магазин", "Хто Замінив", "Username Заміни", "Дата Створення Заявки"
    ]
    try:
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Звіт по замінах"
            sheet.append(headers)
            for cell in sheet[1]:
                style_header_cell(cell)
            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 25
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 20
            sheet.column_dimensions['H'].width = 20
        else:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

        worker_id = replacement_data.get('replacement_worker_id')

        row_data = [
            replacement_data.get('id'),
            replacement_data.get('manager_username'),
            replacement_data.get('request_date'),
            replacement_data.get('position'),
            replacement_data.get('shop'),
            replacement_data.get('replacement_worker_full_name'),
            worker_id,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        sheet.append(row_data)
        workbook.save(filepath)
        logging.info(f"Дані успішно записано у файл: {filepath}")
    except Exception as e:
        logging.error(f"Помилка при записі в Excel: {e}")

if __name__ == '__main__':
    REPORTS_DIR_EXAMPLE = 'reports_example'
    print(f"Назва файлу звіту: {get_report_filename(REPORTS_DIR_EXAMPLE)}")
    test_data = {
        'id': 1,
        'manager_username': 'manager_test',
        'request_date': '30.10.2025',
        'position': 'Пекар',
        'shop': 'Магазин #1 (Центр)',
        'replacement_worker_username': 'worker_test_user'
    }
    record_replacement_to_excel(REPORTS_DIR_EXAMPLE, test_data)
    print("Тестовий запис у файл виконано.")